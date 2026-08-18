"""Lectura de celdas: solo valores, sin ejecutar fórmulas ni objetos."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from nodeone.modules.sales.xls_import.security import XlsImportSecurityError


def load_sheet_grid(kind: str, payload: bytes) -> list[list[Any]]:
    if kind == 'xlsx':
        return _load_xlsx(payload)
    return _load_xls(payload)


def cell_text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def cell_number(value: Any) -> float | None:
    if value is None or value == '':
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = cell_text(value)
    if '%' in text:
        return None
    text = text.replace('B/.', '').replace('$', '').replace('USD', '').strip()
    text = text.replace(' ', '')
    if not text:
        return None
    if text.count(',') == 1 and text.count('.') == 0:
        text = text.replace(',', '.')
    else:
        text = text.replace(',', '')
    try:
        return float(text)
    except ValueError:
        return None


def cell_percent(value: Any) -> float | None:
    """Lee 7%, 0.07 o 7 en contexto de alícuota."""
    if value is None or value == '':
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        if 0 < n <= 1:
            return round(n * 100.0, 4)
        if 0 <= n <= 100:
            return n
        return None
    text = cell_text(value).replace('%', '').replace(',', '.').strip()
    if not text:
        return None
    try:
        n = float(text)
    except ValueError:
        return None
    if 0 < n <= 1:
        return round(n * 100.0, 4)
    if 0 <= n <= 100:
        return n
    return None


def _load_xlsx(payload: bytes) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise XlsImportSecurityError('parser_unavailable', 'Falta openpyxl para leer .xlsx.') from exc
    try:
        try:
            wb = load_workbook(BytesIO(payload), read_only=True, data_only=True, keep_links=False)
        except TypeError:
            wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise XlsImportSecurityError('invalid_content', 'No se pudo leer el .xlsx.') from exc
    try:
        ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            raise XlsImportSecurityError('empty_sheet', 'El libro no tiene hojas.')
        grid: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            grid.append(list(row))
        return grid
    finally:
        wb.close()


def _load_xls(payload: bytes) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise XlsImportSecurityError(
            'parser_unavailable',
            'No se pudo leer el .xls (falta xlrd). Guarde el archivo como .xlsx.',
        ) from exc
    try:
        book = xlrd.open_workbook(file_contents=payload, formatting_info=False)
    except Exception as exc:
        raise XlsImportSecurityError('invalid_content', 'No se pudo leer el .xls.') from exc
    sheet = book.sheet_by_index(0)
    grid: list[list[Any]] = []
    for r in range(sheet.nrows):
        grid.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    return grid
