"""Parser y generador de plantilla XLS para matriz de permisos."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from nodeone.modules.security_matrix_manager.matriz_grid import DEFAULT_MATRIZ_ROWS

SHEET_NAMES = (
    'MATRIZ_GENERAL',
    'GRUPOS_ODOO',
    'USUARIOS',
    'MAPEO_FINAL',
    'PERMISOS_CRITICOS',
)

MATRIZ_HEADERS = ['area', 'modulo', 'pantalla', 'grupo_odoo_xml_id', 'notas']
GRUPOS_HEADERS = ['xml_id', 'nombre', 'categoria']
USUARIOS_HEADERS = ['login', 'nombre', 'activo', 'departamento', 'responsable']
MAPEO_HEADERS = ['login', 'grupo_xml_id', 'accion', 'notas']
CRITICOS_HEADERS = ['grupo_xml_id', 'nombre', 'nivel_riesgo', 'notas']


def _norm_header(cell: Any) -> str:
    return str(cell or '').strip().lower().replace(' ', '_')


def _sheet_rows(ws: Worksheet) -> list[dict[str, Any]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [_norm_header(c) for c in header_row]
    out: list[dict[str, Any]] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        data['_row_number'] = row_num
        out.append(data)
    return out


def parse_workbook_bytes(data: bytes) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parsed: dict[str, list[dict[str, Any]]] = {}
    for name in SHEET_NAMES:
        if name in wb.sheetnames:
            parsed[name] = _sheet_rows(wb[name])
        else:
            parsed[name] = []
    wb.close()
    return parsed


def build_template_bytes(catalog: dict[str, Any]) -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws_matriz = wb.create_sheet('MATRIZ_GENERAL')
    ws_matriz.append(MATRIZ_HEADERS)
    for row in DEFAULT_MATRIZ_ROWS:
        ws_matriz.append([
            row.get('area') or '',
            row.get('modulo') or '',
            row.get('pantalla') or '',
            row.get('grupo_odoo_xml_id') or '',
            '',
        ])

    ws_grupos = wb.create_sheet('GRUPOS_ODOO')
    ws_grupos.append(GRUPOS_HEADERS)
    for g in catalog.get('groups') or []:
        ws_grupos.append([
            g.get('xml_id') or '',
            g.get('name') or '',
            g.get('category') or '',
        ])

    ws_users = wb.create_sheet('USUARIOS')
    ws_users.append(USUARIOS_HEADERS)
    for u in catalog.get('users') or []:
        if not u.get('active', True):
            continue
        ws_users.append([
            u.get('login') or '',
            u.get('name') or '',
            'si' if u.get('active') else 'no',
            u.get('department') or '',
            u.get('parent_login') or '',
        ])

    ws_map = wb.create_sheet('MAPEO_FINAL')
    ws_map.append(MAPEO_HEADERS)
    ws_map.append(['usuario@empresa.com', 'base.group_user', 'add', 'Ejemplo agregar'])

    ws_crit = wb.create_sheet('PERMISOS_CRITICOS')
    ws_crit.append(CRITICOS_HEADERS)
    for cg in catalog.get('critical_groups') or []:
        ws_crit.append([
            cg.get('xml_id') or '',
            cg.get('name') or '',
            cg.get('risk_level') or 'critical',
            cg.get('label') or '',
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pick(row: dict[str, Any], *keys: str) -> str:
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ''


def normalize_mapeo_action(raw: str) -> str | None:
    s = (raw or '').strip().lower()
    if s in ('add', 'agregar', 'añadir', 'alta', 'si', 'sí', 'yes', '1', 'true'):
        return 'add'
    if s in ('remove', 'quitar', 'baja', 'no', '0', 'false', 'eliminar'):
        return 'remove'
    return None
