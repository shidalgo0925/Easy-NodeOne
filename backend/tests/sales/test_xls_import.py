"""Tests — importador XLS de Ventas (parser, recálculo, seguridad). No toca FE."""

from __future__ import annotations

import io
import sys
import unittest
import zipfile
from datetime import date
from pathlib import Path

backend_dir = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(backend_dir))


def build_factura_xlsx(number: int = 496, *, declared_total: float | None = None) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Factura'
    ws['A1'] = 'FACTURA'
    ws['B1'] = number
    ws['A2'] = 'FECHA'
    ws['B2'] = date(2026, 8, 1)
    ws['A3'] = 'CLIENTE'
    ws['B3'] = 'ACME COMERCIAL SA'
    ws['A4'] = 'RUC'
    ws['B4'] = '15561234-2-2023'
    ws['A5'] = 'DV'
    ws['B5'] = '15'
    ws['A6'] = 'DIRECCION'
    ws['B6'] = 'Calle 50, Panamá'
    ws['A7'] = 'TELEFONO'
    ws['B7'] = '507-123-4567'
    ws['A8'] = 'EMAIL'
    ws['B8'] = 'facturas@acme.test'
    ws.append([])
    ws.append(['DESCRIPCION', 'CANTIDAD', 'PRECIO', 'ITBMS', 'TOTAL'])
    ws.append(['Consultoría técnica', 2, 100.0, 7, 214.0])
    ws.append(['SUBTOTAL', None, None, None, 200.0])
    ws.append(['ITBMS', None, None, None, 14.0])
    ws.append(['TOTAL', None, None, None, 214.0 if declared_total is None else declared_total])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_blunt_factura_xlsx(number: int = 496) -> bytes:
    """Layout visual tipo FACTURA 496 (Detalles / P. Unitario / P. Total)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Factura'
    ws['A1'] = 'BLUNT CIGARS, S.A.'
    ws['A2'] = f'Factura {number:05d}'
    ws['A3'] = 'Facturado el: 29/07/2026'
    ws['A5'] = 'NOMBRE:'
    ws['B5'] = 'G.M. HOLDING, INC'
    ws['A6'] = 'RUC:'
    ws['B6'] = '0000000000'
    ws['A7'] = 'TELEFONO:'
    ws['A8'] = 'Direccion:'
    ws['B8'] = 'MEDELLIN COLOMBIA'
    ws.append([])
    ws.append(['Detalles', 'Cantidad', 'P.', 'P.'])
    ws.append([None, None, 'Unitario', 'Total'])
    ws.append(['BW PLATINUM 2X -PAPEL HOMOGENIZADO', 845, '$ 266.88', '$ 225,509.38'])
    ws.append(['BW HEMP 4X - PAPEL HOMOGENIZADO', 200, '$ 187.00', '$ 37,400.00'])
    ws.append(['DMCE', 0, '$ 25.00', '$ 25.00'])
    ws.append(['Total', 1045, None, None])
    ws.append(['Sub Total', None, None, 262934.38])
    ws.append(['ITBMS', '7%', None, None])
    ws.append(['Total a Pagar', None, None, 262934.38])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSecurity(unittest.TestCase):
    def test_rejects_csv_and_xlsm(self):
        from nodeone.modules.sales.xls_import.security import XlsImportSecurityError, inspect_upload

        with self.assertRaises(XlsImportSecurityError):
            inspect_upload(filename='a.csv', data=b'a,b\n1,2\n')
        with self.assertRaises(XlsImportSecurityError):
            inspect_upload(filename='macro.xlsm', data=b'PK\x03\x04junk')

    def test_rejects_path_traversal_name_and_accepts_xlsx(self):
        from nodeone.modules.sales.xls_import.security import inspect_upload, sanitize_filename

        self.assertEqual(sanitize_filename('../../etc/passwd.xlsx'), 'passwd.xlsx')
        data = build_factura_xlsx(496)
        wb = inspect_upload(filename='FACTURA 496.xlsx', data=data)
        self.assertEqual(wb.kind, 'xlsx')
        self.assertEqual(len(wb.sha256), 64)

    def test_rejects_vba_in_xlsx_zip(self):
        from nodeone.modules.sales.xls_import.security import XlsImportSecurityError, inspect_upload

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w') as zf:
            zf.writestr('[Content_Types].xml', '<Types/>')
            zf.writestr('xl/vbaProject.bin', b'MZ')
        with self.assertRaises(XlsImportSecurityError) as ctx:
            inspect_upload(filename='ok.xlsx', data=buf.getvalue())
        self.assertEqual(ctx.exception.code, 'macros_not_allowed')


class TestPanamaProfile(unittest.TestCase):
    def test_parses_factura_496_and_497_without_hardcoded_number(self):
        from nodeone.modules.sales.xls_import.detect import parse_grid
        from nodeone.modules.sales.xls_import.security import inspect_upload
        from nodeone.modules.sales.xls_import.workbook import load_sheet_grid

        for num in (496, 497, 498):
            raw = build_factura_xlsx(num)
            wb = inspect_upload(filename=f'FACTURA {num}.xlsx', data=raw)
            grid = load_sheet_grid(wb.kind, wb.payload)
            data, code, ver = parse_grid(wb.filename, grid, 'auto')
            self.assertEqual(code, 'panama_factura_layout')
            self.assertEqual(ver, 1)
            self.assertEqual(data.external_number, str(num))
            self.assertEqual(data.customer, 'ACME COMERCIAL SA')
            self.assertEqual(data.tax_id, '15561234-2-2023')
            self.assertEqual(data.dv, '15')
            self.assertEqual(len(data.lines), 1)
            self.assertEqual(data.lines[0].description, 'Consultoría técnica')
            self.assertEqual(data.lines[0].quantity, 2)
            self.assertEqual(data.lines[0].unit_price, 100.0)
            self.assertEqual(data.lines[0].tax_rate, 7)
            self.assertEqual(data.declared_total, 214.0)


    def test_parses_blunt_visual_layout(self):
        from nodeone.modules.sales.xls_import.detect import parse_grid
        from nodeone.modules.sales.xls_import.security import inspect_upload
        from nodeone.modules.sales.xls_import.totals import recompute
        from nodeone.modules.sales.xls_import.workbook import load_sheet_grid

        raw = build_blunt_factura_xlsx(496)
        wb = inspect_upload(filename='FACTURA 496.xlsx', data=raw)
        grid = load_sheet_grid(wb.kind, wb.payload)
        data, code, ver = parse_grid(wb.filename, grid, 'auto')
        self.assertEqual(code, 'panama_factura_layout')
        self.assertEqual(ver, 1)
        self.assertEqual(data.external_number, '496')
        self.assertEqual(data.date, '2026-07-29')
        self.assertEqual(data.customer, 'G.M. HOLDING, INC')
        self.assertIsNone(data.tax_id)
        self.assertEqual(data.address, 'MEDELLIN COLOMBIA')
        self.assertEqual(len(data.lines), 3)
        self.assertEqual(data.lines[0].description, 'BW PLATINUM 2X -PAPEL HOMOGENIZADO')
        self.assertEqual(data.lines[0].quantity, 845)
        self.assertEqual(data.lines[2].description, 'DMCE')
        self.assertEqual(data.lines[2].quantity, 1.0)
        self.assertEqual(data.declared_subtotal, 262934.38)
        self.assertEqual(data.declared_total, 262934.38)
        self.assertEqual(data.declared_tax, 0.0)
        totals = recompute(data)
        self.assertTrue(totals.within_tolerance, totals.errors)
        self.assertAlmostEqual(totals.grand_total, 262934.38, places=2)
        self.assertFalse(totals.errors)


class TestTotals(unittest.TestCase):
    def test_recompute_matches_declared(self):
        from nodeone.modules.sales.xls_import.detect import parse_grid
        from nodeone.modules.sales.xls_import.security import inspect_upload
        from nodeone.modules.sales.xls_import.totals import recompute
        from nodeone.modules.sales.xls_import.workbook import load_sheet_grid

        raw = build_factura_xlsx(496)
        wb = inspect_upload(filename='FACTURA 496.xlsx', data=raw)
        grid = load_sheet_grid(wb.kind, wb.payload)
        data, _, _ = parse_grid(wb.filename, grid)
        totals = recompute(data)
        self.assertEqual(totals.subtotal, 200.0)
        self.assertEqual(totals.tax_total, 14.0)
        self.assertEqual(totals.grand_total, 214.0)
        self.assertTrue(totals.within_tolerance)
        self.assertFalse(totals.errors)

    def test_mismatch_blocks(self):
        from nodeone.modules.sales.xls_import.detect import parse_grid
        from nodeone.modules.sales.xls_import.security import inspect_upload
        from nodeone.modules.sales.xls_import.totals import recompute
        from nodeone.modules.sales.xls_import.workbook import load_sheet_grid

        raw = build_factura_xlsx(496, declared_total=999.0)
        wb = inspect_upload(filename='FACTURA 496.xlsx', data=raw)
        grid = load_sheet_grid(wb.kind, wb.payload)
        data, _, _ = parse_grid(wb.filename, grid)
        totals = recompute(data)
        self.assertFalse(totals.within_tolerance)
        self.assertTrue(any('no coincide' in e.lower() for e in totals.errors))


class TestStorageRoot(unittest.TestCase):
    def test_storage_root_honors_env(self):
        import os
        import tempfile

        from nodeone.modules.sales.xls_import import service as svc

        with tempfile.TemporaryDirectory() as tmp:
            prev = os.environ.get('NODEONE_SALES_XLS_STORAGE')
            os.environ['NODEONE_SALES_XLS_STORAGE'] = tmp
            try:
                svc._STORAGE_ROOT_CACHE = None
                root = svc._storage_root()
                self.assertEqual(Path(root).resolve(), Path(tmp).resolve())
            finally:
                svc._STORAGE_ROOT_CACHE = None
                if prev is None:
                    os.environ.pop('NODEONE_SALES_XLS_STORAGE', None)
                else:
                    os.environ['NODEONE_SALES_XLS_STORAGE'] = prev

    def test_store_file_writes_when_env_set(self):
        import os
        import tempfile

        from nodeone.modules.sales.xls_import import service as svc
        from nodeone.modules.sales.xls_import.security import inspect_upload

        raw = build_blunt_factura_xlsx(496)
        wb = inspect_upload(filename='FACTURA 496.xlsx', data=raw)
        with tempfile.TemporaryDirectory() as tmp:
            prev = os.environ.get('NODEONE_SALES_XLS_STORAGE')
            os.environ['NODEONE_SALES_XLS_STORAGE'] = tmp
            try:
                svc._STORAGE_ROOT_CACHE = None
                stored = svc._store_file(12, wb)
                path = Path(tmp) / stored
                self.assertTrue(path.is_file())
                self.assertGreater(path.stat().st_size, 0)
            finally:
                svc._STORAGE_ROOT_CACHE = None
                if prev is None:
                    os.environ.pop('NODEONE_SALES_XLS_STORAGE', None)
                else:
                    os.environ['NODEONE_SALES_XLS_STORAGE'] = prev


class TestCatalogMatchHelpers(unittest.TestCase):
    def test_catalog_key_collapses_punctuation(self):
        from nodeone.modules.sales.xls_import.service import _catalog_key

        self.assertEqual(_catalog_key('G.M. HOLDING, INC'), 'gmholdinginc')
        self.assertEqual(_catalog_key('GM HOLDING INC'), 'gmholdinginc')
        self.assertEqual(_catalog_key('BW PLATINUM 2X -PAPEL HOMOGENIZADO'), _catalog_key('bw platinum 2x papel homogenizado'))

    def test_usable_tax_id_drops_placeholders(self):
        from nodeone.modules.sales.xls_import.service import _usable_tax_id

        self.assertEqual(_usable_tax_id('0000000000'), '')
        self.assertEqual(_usable_tax_id('--'), '')
        self.assertEqual(_usable_tax_id('15561234-2-2023'), '15561234-2-2023')


class TestRoutesRegistered(unittest.TestCase):
    def test_endpoints_exist(self):
        from app import app

        eps = {r.endpoint for r in app.url_map.iter_rules()}
        self.assertIn('sales.xls_import_analyze', eps)
        self.assertIn('sales.xls_import_commit', eps)
        self.assertIn('admin_sales_xls_import', eps)

    def test_confirm_still_separate_from_fe(self):
        from app import app

        rules = {str(r.rule) for r in app.url_map.iter_rules()}
        self.assertIn('/api/sales/quotations/<int:qid>/confirm', rules)
        self.assertIn('/api/sales/quotations/<int:qid>/create-invoice', rules)
        self.assertIn('/api/admin/efactura/issue-from-invoice/<int:invoice_id>', rules)


if __name__ == '__main__':
    unittest.main()
